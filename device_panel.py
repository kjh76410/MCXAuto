import os
import re
import time
import json
import datetime
import threading
import subprocess
import importlib

from PySide6.QtCore import Qt, QTimer, QSize
from PySide6.QtGui import (
    QColor,
    QFontMetrics,
    QIcon,
    QIntValidator,
    QTextCursor,
    QTextCharFormat,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QButtonGroup,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
import qtawesome as qta

import adb_logic
from file_manager import FileManager
from ui_common import (
    Palette,
    kfont,
    styled,
    add_shadow,
    card_css,
    btn_css,
    make_button,
    clear_layout,
    Signals,
    QtLogConsole,
    ClickableLabel,
    SegmentedButton,
    PulseCanvas,
)


class DevicePanel(QWidget):
    """단말기 한 대를 담당하는 패널. 미러링/그룹·유저 리스트/SIP Flow/로그 등
    기존에 App(단일 기기 전용)이 갖고 있던 상태와 위젯을 그대로 옮겨왔습니다 - 두 대를
    동시에 띄우기 위해 이 패널을 두 번 인스턴스화합니다(App.panel_a / App.panel_b).
    기기 탐색(adb devices 스캔)만은 App 쪽에 남아있고, 이 패널은 device_ready 시그널로
    결과를 받아 반영하기만 합니다.

    Test Results 표는 한 벌의 TC 목록을 두 단말이 같이 채우는 게 자연스러워서
    (엑셀로 가져오는 TC 목록이 원래 하나뿐이므로) 패널마다 따로 두지 않고, App이 만든
    단 하나의 ResultsPanel 인스턴스를 두 DevicePanel이 공유해서 씁니다."""

    def __init__(self, panel_label="A", results_panel=None, parent=None):
        super().__init__(parent)
        self.panel_label = panel_label
        # 두 패널이 공유하는 단 하나의 Test Results 표 (App이 만들어서 넘겨줍니다).
        self.results_panel = results_panel

        self.current_uuid = None
        self.is_log_on = False
        self.is_pcap_on = False
        self.has_private_call = False
        self.project_name = "알 수 없는 프로젝트"
        self.current_mode = "call"
        self.stop_event = threading.Event()

        self._realtime_sip_stop_event = None
        self._realtime_sip_state = None
        self._realtime_sip_thread = None
        self._net_monitor_gen = 0
        self._sip_analyzer_gen = 0
        self._flow_dedupe = {}
        self._pending_dnd_reason = False
        self._sip_log_process = None
        self._pulse_idle_timer = None
        self.log_proc = None
        self.log_file = None

        self.group_check_vars = {}
        self.all_cards = []
        self.user_ui_registry = {}

        self.log_console = QtLogConsole(self)

        self.signals = Signals()
        self.signals.log_append.connect(self._append_log)
        self.signals.flow_card.connect(self._add_flow_card_ui)
        self.signals.network_label.connect(self._update_network_label)
        self.signals.floor_state.connect(self._on_floor_state)
        self.signals.device_ready.connect(self._on_device_ready)
        self.signals.pcap_state.connect(self._set_pcap_ui_state)

        self._build_panel_ui()
        QTimer.singleShot(300, self.pulse_canvas.stop)

    # ==========================================
    # 레이아웃 구성
    # ==========================================
    def _build_panel_ui(self):
        """이 패널(A 또는 B) 하나만으로 자기 완결적인 위젯 트리를 만드는 대신, 조각(상단
        헤더/배너, 미러링+리스트 컬럼, 로그카드 컬럼)만 만들어 self.top_block /
        self.left_column_widget / self.right_column_widget 로 들고 있습니다. 실제 화면
        배치(그리드에 두 패널을 나란히 놓고, 로그카드 두 개 아래에 결과표를 하나만 병합해
        붙이는 것)는 App(ui_logic.py)이 담당합니다."""
        self.top_block = QWidget()
        top_layout = QVBoxLayout(self.top_block)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(10)
        top_layout.addWidget(self._build_header())
        top_layout.addWidget(self._build_top_banner())

        self.left_column_widget = self._build_left_column()
        self.right_column_widget = self._build_right_column()

    def _build_header(self):
        """예전 좌측 사이드바(280px 고정 세로 컬럼)를 두 패널이 나란히 들어갈 수 있도록
        압축한 상단 가로 헤더로 바꾼 버전입니다. 환경/WiFi/앱 설치·삭제/시나리오 실행 같은
        부차 기능은 ⚙ 관리 메뉴 하나로 몰아넣었습니다."""
        frame = styled(QFrame(), f"background-color:{Palette.bg}; border-radius:5px;")
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(10, 8, 10, 8)
        layout.setSpacing(3)

        top_row = QHBoxLayout()
        top_row.setSpacing(8)
        self.lbl_project = QLabel(f"[{self.panel_label}] 프로젝트: 대기 중")
        self.lbl_project.setFont(kfont(13, True))
        self.lbl_project.setStyleSheet(f"color:{Palette.blue};")
        top_row.addWidget(self.lbl_project)
        top_row.addStretch(1)
        self.lbl_network = QLabel("네트워크: -")
        self.lbl_network.setFont(kfont(10))
        self.lbl_network.setStyleSheet(f"color:{Palette.text_main};")
        top_row.addWidget(self.lbl_network)
        layout.addLayout(top_row)

        self.label = QLabel("단말을 연결해주세요.")
        self.label.setFont(kfont(10))
        self.label.setStyleSheet(f"color:{Palette.text_sub};")
        layout.addWidget(self.label)

        info_row = QHBoxLayout()
        info_row.setSpacing(10)
        self.lbl_model = QLabel("모델: -")
        self.lbl_hw_version = QLabel("HW: -")
        self.lbl_android_ver = QLabel("Android: -")
        self.lbl_os_build = QLabel("OS: -")
        self.lbl_version = QLabel("버전: -")
        self.lbl_project_version = self.lbl_version  # 하위 호환: 예전 이름으로도 접근 가능
        for lbl in (self.lbl_model, self.lbl_hw_version, self.lbl_android_ver, self.lbl_os_build, self.lbl_version):
            lbl.setFont(kfont(9))
            lbl.setStyleSheet(f"color:{Palette.text_main};")
            info_row.addWidget(lbl)
        info_row.addStretch(1)
        layout.addLayout(info_row)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self.btn_connect = self._make_button(
            "기기 연결", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover,
            height=28, icon_name="fa5s.plug", icon_size=13,
        )
        self.btn_manage = self._make_button(
            "⚙ 관리", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover,
            height=28, icon_name="fa5s.tools", icon_size=13,
        )
        self.btn_manage.clicked.connect(self._open_manage_menu)
        btn_row.addWidget(self.btn_connect, 1)
        btn_row.addWidget(self.btn_manage, 1)
        layout.addLayout(btn_row)

        return frame

    def _open_manage_menu(self):
        """환경/WiFi 설정, 앱 설치/삭제, 시나리오 실행 같은 부차 기능들을 한 버튼 아래로 모은 메뉴."""
        menu = QMenu(self)
        menu.addAction("환경 설정", self.open_env_setup)
        menu.addAction("WiFi 설정", self.open_wifi_setup)
        menu.addSeparator()
        menu.addAction("앱 설치 (.apk)", self.run_install_app)
        menu.addAction("데이터 삭제", self.run_clear_data)
        menu.addAction("앱 삭제", self.run_uninstall_app)
        menu.addSeparator()
        menu.addAction("전체 시나리오 실행", self.run_automation)
        menu.addAction("중지", self.stop_automation)
        menu.addAction("단위 테스트", self.open_unit_test_popup)
        menu.exec(self.btn_manage.mapToGlobal(self.btn_manage.rect().bottomLeft()))

    def _build_top_banner(self):
        banner = styled(QFrame(), card_css())
        banner.setFixedHeight(40)
        layout = QHBoxLayout(banner)
        layout.setContentsMargins(12, 4, 12, 4)
        self.feature_tag_frame = QWidget()
        self.feature_tag_layout = QHBoxLayout(self.feature_tag_frame)
        self.feature_tag_layout.setContentsMargins(0, 0, 0, 0)
        self.feature_tag_layout.setSpacing(0)
        layout.addWidget(self.feature_tag_frame)
        layout.addStretch(1)
        self._reset_feature_tags()
        return add_shadow(banner)

    # ---------- 왼쪽: 미러링(위) + Group/User List(아래) ----------
    def _build_left_column(self):
        self.phone_width = 190
        self.phone_height = 290

        col = QWidget()
        col_layout = QVBoxLayout(col)
        col_layout.setContentsMargins(0, 0, 0, 0)
        col_layout.setSpacing(10)
        col.setFixedWidth(300)

        # 🔥 미러링 카드는 컬럼 폭에 맞춰 늘리지 않고 내용물(작은 화면) 크기만큼만 차지하도록
        # AlignHCenter로 추가합니다. 그래야 카드 배경이 남는 여백 없이 딱 붙습니다.
        col_layout.addWidget(self._build_mirror_card(), 0, Qt.AlignHCenter)
        col_layout.addWidget(self._build_list_card(), 1)
        return col

    def _build_mirror_card(self):
        card = styled(QFrame(), card_css())
        outer = QVBoxLayout(card)
        outer.setContentsMargins(14, 8, 14, 8)
        outer.setSpacing(4)
        outer.setAlignment(Qt.AlignHCenter)

        btn_kwargs = dict(bg=Palette.neutral_bg, fg=Palette.text_main, hover=Palette.neutral_hover, height=25, radius=5)

        top_nav = QHBoxLayout()
        top_nav.setSpacing(2)
        btn_capture = self._make_button("캡쳐", icon_name="fa5s.camera", **btn_kwargs)
        btn_capture.clicked.connect(self.capture_screen)
        btn_record = self._make_button("촬영", icon_name="fa5s.video", **btn_kwargs)
        btn_record.clicked.connect(self.record_screen)
        top_nav.addWidget(btn_capture)
        top_nav.addWidget(btn_record)
        outer.addLayout(top_nav)

        self.mirror_container = styled(QWidget(), "background-color:#1C1C1E;")
        self.mirror_container.setFixedSize(self.phone_width, self.phone_height)
        self.mirror_container.setAttribute(Qt.WA_NativeWindow, True)
        self.lbl_placeholder = QLabel("대기 중", self.mirror_container)
        self.lbl_placeholder.setGeometry(0, 0, self.phone_width, self.phone_height)
        self.lbl_placeholder.setAlignment(Qt.AlignCenter)
        self.lbl_placeholder.setStyleSheet(f"color:{Palette.text_sub}; background:transparent;")
        self.lbl_placeholder.setFont(kfont(11))
        outer.addWidget(self.mirror_container)

        bottom_nav = QHBoxLayout()
        bottom_nav.setSpacing(2)
        btn_back = self._make_button("", icon_name="fa5s.chevron-left", icon_size=16, **btn_kwargs)
        btn_back.clicked.connect(lambda: self.send_adb_keyevent(4))
        btn_home = self._make_button("", icon_name="fa5s.circle", icon_size=16, **btn_kwargs)
        btn_home.clicked.connect(lambda: self.send_adb_keyevent(3))
        btn_recent = self._make_button("", icon_name="fa5s.square", icon_size=16, **btn_kwargs)
        btn_recent.clicked.connect(lambda: self.send_adb_keyevent(187))
        bottom_nav.addWidget(btn_recent)
        bottom_nav.addWidget(btn_home)
        bottom_nav.addWidget(btn_back)
        outer.addLayout(bottom_nav)

        return add_shadow(card)

    def _build_list_card(self):
        card = styled(QFrame(), card_css())
        layout = QVBoxLayout(card)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        header = QHBoxLayout()
        self.btn_tab_group = self._make_button("Group List", Palette.neutral_hover, Palette.text_main, Palette.neutral_hover, height=28, radius=6)
        self.btn_tab_group.clicked.connect(lambda: self.switch_tab("group"))
        self.btn_tab_user = self._make_button("User List", Palette.neutral_bg, Palette.text_sub, Palette.neutral_hover, height=28, radius=6)
        self.btn_tab_user.clicked.connect(lambda: self.switch_tab("user"))
        self.btn_refresh = self._make_button("", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, height=28, radius=6, icon_name="fa5s.sync-alt")
        self.btn_refresh.setFixedWidth(34)
        self.btn_refresh.clicked.connect(self.refresh_all_lists)
        header.addWidget(self.btn_tab_group, 1)
        header.addWidget(self.btn_tab_user, 1)
        header.addWidget(self.btn_refresh)
        layout.addLayout(header)

        mode_row = QHBoxLayout()
        lbl_mode = QLabel("테스트 모드:")
        lbl_mode.setFont(kfont(11))
        lbl_mode.setStyleSheet(f"color:{Palette.text_sub};")
        self.seg_mode_toggle = SegmentedButton(["📞 통화", "💬 메시지"], selected_color=Palette.neutral_hover, height=25, font=kfont(11, True))
        self.seg_mode_toggle.set("📞 통화")
        self.seg_mode_toggle.changed.connect(self.on_mode_toggle_changed)
        mode_row.addWidget(lbl_mode)
        mode_row.addWidget(self.seg_mode_toggle, 1)
        layout.addLayout(mode_row)

        self.list_stack = QStackedWidget()
        self.group_scroll, self.group_list_container, self.group_list_layout = self._make_scroll_list()
        self.user_scroll, self.user_list_container, self.user_list_layout = self._make_scroll_list()
        self.list_stack.addWidget(self.group_scroll)
        self.list_stack.addWidget(self.user_scroll)
        layout.addWidget(self.list_stack, 1)

        self.my_id_label = QLabel("내 정보: 연결 대기")
        self.my_id_label.setFont(kfont(11, True))
        self.my_id_label.setStyleSheet(f"color:{Palette.blue};")
        layout.addWidget(self.my_id_label)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(4)
        self.btn_group_call = self._make_button("통화 발신", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, height=34, icon_name="fa5s.phone")
        self.btn_group_call.clicked.connect(self.on_main_call_button_clicked)
        self.btn_group_msg = self._make_button("메시지 전송", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, height=34, icon_name="fa5s.comment-dots")
        self.btn_group_msg.clicked.connect(self.send_group_message)
        btn_row.addWidget(self.btn_group_call)
        btn_row.addWidget(self.btn_group_msg)
        layout.addLayout(btn_row)

        return add_shadow(card)

    def _make_scroll_list(self):
        area = QScrollArea()
        area.setWidgetResizable(True)
        area.setFrameShape(QFrame.NoFrame)
        area.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        container = styled(QWidget(), "background: transparent;")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)
        layout.addStretch(1)
        area.setWidget(container)
        return area, container, layout

    # ---------- 오른쪽: PTT Pulse + Network & System Logs ----------
    def _build_right_column(self):
        col = QWidget()
        layout = QVBoxLayout(col)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        # PTT Floor State 패널은 제거되고 로그 화면이 그 자리를 채웁니다.
        # 발언권 상태를 추적하던 위젯들은 화면에는 그리지 않고 내부 상태 갱신용으로만 유지합니다.
        self.lbl_pulse_status = QLabel("대기")
        self.pulse_canvas = PulseCanvas(color=Palette.blue)

        monitor = styled(QFrame(), card_css())
        monitor_layout = QVBoxLayout(monitor)
        monitor_layout.setContentsMargins(12, 10, 12, 12)
        monitor_layout.setSpacing(4)

        monitor_header = QHBoxLayout()
        lbl_title = QLabel("Network & System Logs")
        lbl_title.setFont(kfont(12, True))
        lbl_title.setStyleSheet(f"color:{Palette.text_main};")
        monitor_header.addWidget(lbl_title)
        monitor_header.addStretch(1)

        self.btn_toggle_log = self._make_button("Logcat ON", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, height=24, radius=5, icon_name="fa5s.file-alt")
        self.btn_toggle_log.setFixedWidth(105)
        self.btn_toggle_log.clicked.connect(self.toggle_log)
        self.btn_toggle_pcap = self._make_button("PCAP ON", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, height=24, radius=5, icon_name="fa5s.circle", icon_size=10)
        self.btn_toggle_pcap.setFixedWidth(95)
        self.btn_toggle_pcap.clicked.connect(self.toggle_pcap)
        monitor_header.addWidget(self.btn_toggle_log)
        monitor_header.addWidget(self.btn_toggle_pcap)
        monitor_layout.addLayout(monitor_header)

        self.tab_view = QTabWidget()
        self.tab_view.setStyleSheet(
            f"QTabBar::tab {{ background:{Palette.neutral_bg}; padding:5px 12px; border-radius:3px; margin:3px 2px; }}"
            f"QTabBar::tab:selected {{ background:{Palette.blue}; color:white; }}"
            f"QTabWidget::pane {{ border:none; }}"
        )

        sip_tab = QWidget()
        sip_layout = QVBoxLayout(sip_tab)
        sip_layout.setContentsMargins(4, 4, 4, 4)
        self.flow_scroll, self.flow_container, self.flow_layout = self._make_scroll_list()
        sip_layout.addWidget(self.flow_scroll)
        self.tab_view.addTab(sip_tab, "SIP Flow")
        self._show_sip_placeholder()

        log_tab = QWidget()
        log_layout = QVBoxLayout(log_tab)
        log_layout.setContentsMargins(4, 4, 4, 4)
        log_layout.setSpacing(6)
        self.entry_search = QLineEdit()
        self.entry_search.setPlaceholderText("🔍 터미널 로그 실시간 검색")
        self.entry_search.setFixedHeight(30)
        self.entry_search.setStyleSheet(
            f"QLineEdit {{ background-color:white; border:1px solid {Palette.border}; border-radius:4px; padding:0 10px; }}"
        )
        self.entry_search.textChanged.connect(self._filter_log)
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setFont(kfont(10))
        self.txt_log.setStyleSheet(
            f"QTextEdit {{ background-color:{Palette.bg}; color:{Palette.text_main}; "
            f"border:1px solid {Palette.border}; border-radius:4px; }}"
        )
        log_layout.addWidget(self.entry_search)
        log_layout.addWidget(self.txt_log, 1)
        self.tab_view.addTab(log_tab, "System Log")

        monitor_layout.addWidget(self.tab_view, 1)
        layout.addWidget(add_shadow(monitor), 1)

        self.txt_log.append("[Terminal] 시스템 로그 출력을 대기 중입니다...")
        return col

    # ==========================================
    # 🧩 공용 UI 헬퍼
    # ==========================================
    def _section_label(self, text):
        lbl = QLabel(text)
        lbl.setFont(kfont(12, True))
        lbl.setStyleSheet(f"color:{Palette.text_main};")
        return lbl

    def _hline(self):
        line = QFrame()
        line.setFixedHeight(1)
        line.setStyleSheet(f"background-color:{Palette.border};")
        return line

    def _make_button(self, text, bg, fg, hover, height=26, radius=Palette.radius, icon_name=None, icon_size=14):
        return make_button(text, bg, fg, hover, height, radius, icon_name, icon_size)

    def _show_sip_placeholder(self):
        clear_layout(self.flow_layout, keep=0)
        self.flow_layout.addStretch(1)
        lbl = QLabel("단말을 연결하면 실시간 SIP/Call Flow가 이곳에 표시됩니다.")
        lbl.setFont(kfont(12))
        lbl.setStyleSheet(f"color:{Palette.text_sub};")
        lbl.setAlignment(Qt.AlignCenter)
        self.flow_layout.insertWidget(0, lbl)

    def _reset_feature_tags(self):
        clear_layout(self.feature_tag_layout, keep=0)
        lbl = QLabel("단말기를 연결하면 이곳에 프로젝트 지원 기능이 표시됩니다.")
        lbl.setFont(kfont(11))
        lbl.setStyleSheet(f"color:{Palette.text_sub};")
        self.feature_tag_layout.addWidget(lbl)

    def _filter_log(self, text):
        # 검색어 입력 시 하이라이트 대신 커서 위치의 다음 일치 항목으로 이동
        if not text:
            return
        cursor = self.txt_log.textCursor()
        found = self.txt_log.find(text)
        if not found:
            cursor.movePosition(QTextCursor.Start)
            self.txt_log.setTextCursor(cursor)
            self.txt_log.find(text)

    # ==========================================
    # 유틸리티: 스레드 안전 로그/카드 업데이트
    # ==========================================
    def safe_log_insert(self, text, is_error=False):
        self.signals.log_append.emit(text, is_error)

    def _append_log(self, text, is_error):
        cursor = self.txt_log.textCursor()
        cursor.movePosition(QTextCursor.End)
        fmt = QTextCharFormat()
        fmt.setForeground(QColor(Palette.danger if is_error else Palette.text_main))
        cursor.setCharFormat(fmt)
        cursor.insertText(text if text.endswith("\n") else text + "\n")
        self.txt_log.setTextCursor(cursor)
        self.txt_log.ensureCursorVisible()

    def add_flow_card(self, event_type, title, detail, is_error=False):
        self.signals.flow_card.emit(event_type, title, detail, is_error)

    def _add_flow_card_ui(self, event_type, title, detail, is_error):
        if is_error:
            b_color, bg_col, badge_text = Palette.danger, Palette.danger_bg, "ERROR"
        elif event_type == "RX":
            b_color, bg_col, badge_text = Palette.blue, Palette.tint_blue_bg, "RECV"
        else:
            b_color, bg_col, badge_text = Palette.blue, Palette.panel, "PROC"

        card = styled(QFrame(), f"background-color:{bg_col}; border:none; border-radius:4px;")
        row = QHBoxLayout(card)
        row.setContentsMargins(10, 6, 10, 6)
        row.setSpacing(8)

        badge = QLabel(badge_text)
        badge.setFont(kfont(9, True))
        badge.setFixedHeight(16)
        badge.setAlignment(Qt.AlignCenter)
        badge.setStyleSheet(f"background-color:{b_color}; color:white; border-radius:3px; padding:0 6px;")
        lbl_title = QLabel(title)
        lbl_title.setFont(kfont(12, True))
        lbl_title.setStyleSheet(f"color:{Palette.text_main};")

        lbl_detail = QLabel()
        lbl_detail.setFont(kfont(11))
        lbl_detail.setStyleSheet(f"color:{Palette.danger if is_error else Palette.text_sub};")
        lbl_detail.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        lbl_detail.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        # 상세 내용은 한 행에 다 보이도록 줄바꿈 대신, 폭에 맞춰 말줄임(...)으로 잘라냅니다.
        # 실제 전체 내용은 툴팁으로 확인할 수 있습니다.
        available_w = max(120, self.flow_scroll.viewport().width() - 220)
        elided = QFontMetrics(lbl_detail.font()).elidedText(detail, Qt.ElideRight, available_w)
        lbl_detail.setText(elided)
        lbl_detail.setToolTip(detail)

        row.addWidget(badge, 0, Qt.AlignVCenter)
        row.addWidget(lbl_title, 0, Qt.AlignVCenter)
        row.addWidget(lbl_detail, 1)

        self.flow_layout.insertWidget(self.flow_layout.count() - 1, card)
        QTimer.singleShot(0, lambda: self.flow_scroll.verticalScrollBar().setValue(
            self.flow_scroll.verticalScrollBar().maximum()
        ))

    # ==========================================
    # 🔌 기기 연결 결과 반영 (탐색 자체는 App이 담당)
    # ==========================================
    def _on_device_ready(self, info):
        self.btn_connect.setEnabled(True)

        if info:
            self.current_uuid = info["uuid"]
            model = info["model"]
            android_version = info["android_version"]
            os_build = info["os_build"]
            version_name = info["version_name"]
            hw_version = info["hw_version"]
            self.project_name = FileManager.get_project_name(version_name)

            self.label.setText(f"연결됨: {model}")
            self.label.setStyleSheet(f"color:{Palette.blue};")
            self.lbl_model.setText(f"모델: {model}")
            self.lbl_hw_version.setText(f"HW: {hw_version}")
            self.lbl_android_ver.setText(f"Android: {android_version}")
            self.lbl_os_build.setText(f"OS: {os_build}")
            self.lbl_version.setText(f"버전: {version_name}")
            self.lbl_project.setText(f"[{self.panel_label}] 프로젝트: {self.project_name}")

            self.update_project_features(self.project_name)
            self.refresh_all_lists()

            self.run_mirror()
            self.start_network_monitor()
            self.start_realtime_log_analyzer()
            self.start_realtime_sip_stream()
        else:
            self.current_uuid = None
            self.label.setText("연결된 단말 없음")
            self.label.setStyleSheet(f"color:{Palette.text_sub};")
            self.lbl_model.setText("모델: -")
            self.lbl_hw_version.setText("HW: -")
            self.lbl_android_ver.setText("Android: -")
            self.lbl_os_build.setText("OS: -")
            self.lbl_version.setText("버전: -")
            self.lbl_network.setText("네트워크: -")
            self.lbl_network.setStyleSheet(f"color:{Palette.text_main};")
            self.lbl_project.setText(f"[{self.panel_label}] 프로젝트: 대기 중")
            self.lbl_project.setStyleSheet(f"color:{Palette.blue};")

            self._reset_feature_tags()
            self.pulse_canvas.stop()
            self.lbl_pulse_status.setText("대기")
            self.lbl_pulse_status.setStyleSheet("color:#8E8E93;")

            self.mirror_container.setFixedSize(self.phone_width, self.phone_height)
            self.lbl_placeholder.setGeometry(0, 0, self.phone_width, self.phone_height)
            self.lbl_placeholder.show()

            self.stop_realtime_sip_stream()
            self._sip_analyzer_gen += 1
            if self._sip_log_process:
                try:
                    self._sip_log_process.terminate()
                except Exception:
                    pass
            self._show_sip_placeholder()

    def start_network_monitor(self):
        self._net_monitor_gen += 1
        gen = self._net_monitor_gen
        uuid = self.current_uuid

        def poll():
            while self._net_monitor_gen == gen and self.current_uuid == uuid:
                status = adb_logic.get_network_status(uuid)
                self.signals.network_label.emit(status)
                time.sleep(3)

        threading.Thread(target=poll, daemon=True).start()

    def _update_network_label(self, status):
        if self.current_uuid is None:
            return
        is_down = ("끊김" in status) or ("불가" in status)
        color = Palette.danger if is_down else Palette.blue
        self.lbl_network.setText(f"네트워크: {status}")
        self.lbl_network.setStyleSheet(f"color:{color};")

    def send_adb_keyevent(self, keycode):
        target_device = self.current_uuid

        def run():
            try:
                if target_device:
                    cmd = ["adb", "-s", target_device, "shell", "input", "keyevent", str(keycode)]
                else:
                    cmd = ["adb", "shell", "input", "keyevent", str(keycode)]
                if os.name == "nt":
                    startupinfo = subprocess.STARTUPINFO()
                    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                    subprocess.run(cmd, startupinfo=startupinfo, capture_output=True, text=True)
                else:
                    subprocess.run(cmd, capture_output=True, text=True)
            except Exception as e:
                print(f"⚠️ ADB 명령어 전송 실패: {e}")

        threading.Thread(target=run, daemon=True).start()

    # ==========================================
    # 🎙️ PTT 발언권(Floor) 연동 펄스 애니메이션
    # ==========================================
    def set_ptt_active(self, active: bool):
        if active:
            self.lbl_pulse_status.setText("🔴 송신 중")
            self.lbl_pulse_status.setStyleSheet(f"color:{Palette.danger};")
            self.pulse_canvas.start()
        else:
            self.lbl_pulse_status.setText("대기")
            self.lbl_pulse_status.setStyleSheet("color:#8E8E93;")
            self.pulse_canvas.stop()

    def _on_floor_state(self, state_text):
        """mMBCPKeyEvent(MBCP_UI_EVENT_*) 로그 문자열을 보고 발언권 보유 여부를 추정해 펄스를 연동합니다.
        실측 확인된 값은 MBCP_UI_EVENT_IDLE / MBCP_UI_EVENT_NONE 뿐이라, GRANT 계열 값은 추정치입니다."""
        if self._pulse_idle_timer is not None:
            self._pulse_idle_timer.stop()
            self._pulse_idle_timer = None

        txt = state_text.upper()
        if any(k in txt for k in ("GRANT", "TALK", "TRANSMIT", "TX")):
            self.set_ptt_active(True)
        elif any(k in txt for k in ("IDLE", "NONE", "RELEASE", "DENY", "DENIED", "REVOKE", "STOP")):
            self.set_ptt_active(False)
        else:
            self.set_ptt_active(True)
            self._pulse_idle_timer = QTimer(self)
            self._pulse_idle_timer.setSingleShot(True)
            self._pulse_idle_timer.timeout.connect(lambda: self.set_ptt_active(False))
            self._pulse_idle_timer.start(2000)

    # ==========================================
    # 🏷️ 프로젝트 지원 기능 뱃지
    # ==========================================
    def update_project_features(self, project_name):
        clear_layout(self.feature_tag_layout, keep=0)

        features = FileManager.get_project_features(project_name)
        if not features:
            self.has_private_call = False
            self._reset_feature_tags()
            return

        group_map = {"regroup": "ReGroup", "prearranged": "PreArranged", "chat": "Chat"}
        private_map = {
            "emergency_ptt": "E-PTT",
            "emergency_ptv": "E-PTV",
            "ptt": "PTT",
            "ptv": "PTV",
            "without_floor_control": "W/O Floor",
            "mcvideo_push": "Video Push",
            "mcvideo_pull": "Video Pull",
            "first_answer": "First Answer",
        }
        msg_map = {
            "normal": "일반",
            "emergency": "비상",
            "pre_defined": "Pre-Defined",
            "canned": "상용문구",
            "attachment": "첨부파일",
        }

        private_call_data = features.get("private_call", {})
        self.has_private_call = any(val == 1 for val in private_call_data.values())

        self.btn_tab_user.setEnabled(self.has_private_call)
        if not self.has_private_call and self.current_mode == "user":
            self.switch_tab("group")

        def create_tag_group(title, icon, category_data, name_map, bg_color, txt_color):
            active_features = [name_map[k] for k, v in category_data.items() if v == 1 and k in name_map]
            if not active_features:
                return
            group = QWidget()
            g_layout = QHBoxLayout(group)
            g_layout.setContentsMargins(0, 0, 16, 0)
            g_layout.setSpacing(3)
            lbl = QLabel(f"{icon} {title}:")
            lbl.setFont(kfont(11, True))
            lbl.setStyleSheet(f"color:{Palette.text_sub};")
            g_layout.addWidget(lbl)
            for f_name in active_features:
                badge = QLabel(f_name)
                badge.setFont(kfont(9, True))
                badge.setStyleSheet(
                    f"background-color:{bg_color}; color:{txt_color}; border-radius:5px; padding:2px 8px;"
                )
                badge.setAttribute(Qt.WA_StyledBackground, True)
                g_layout.addWidget(badge)
            self.feature_tag_layout.addWidget(group)

        create_tag_group("Group", "👥", features.get("group_call", {}), group_map, "#E0E7FF", "#4318FF")
        create_tag_group("Private", "👤", private_call_data, private_map, "#DCFCE7", "#05CD99")
        create_tag_group("Message", "✉️", features.get("message", {}), msg_map, "#FFEDD5", "#EE5D50")

        if self.feature_tag_layout.count() == 0:
            self._reset_feature_tags()

    # ==========================================
    # 📋 Group List
    # ==========================================
    def _make_repeat_edit(self):
        edit = QLineEdit("1")
        edit.setFixedSize(40, 24)
        edit.setAlignment(Qt.AlignCenter)
        edit.setValidator(QIntValidator(0, 999))
        edit.setStyleSheet(f"border:1px solid {Palette.border}; border-radius:3px;")

        def on_finish():
            if not edit.text() or edit.text() == "0":
                edit.setText("1")

        edit.editingFinished.connect(on_finish)
        return edit

    def _make_list_card(self, parent_layout, name, id_text, seg_call_values, seg_msg_values):
        """Group/User List에서 공통으로 쓰는 카드(체크박스 + 반복 횟수 + 이름/ID + 통화·메시지 방식) 생성."""
        card = styled(QFrame(), card_css(radius=5))
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(0, 0, 0, 0)
        card_layout.setSpacing(0)

        top_row = QWidget()
        top_layout = QHBoxLayout(top_row)
        top_layout.setContentsMargins(8, 8, 8, 8)
        top_layout.setSpacing(6)

        checkbox = QPushButton()
        checkbox.setCheckable(True)
        checkbox.setFixedSize(22, 22)
        checkbox.setCursor(Qt.PointingHandCursor)
        self._style_checkbox(checkbox)
        top_layout.addWidget(checkbox)

        text_frame = QWidget()
        text_layout = QVBoxLayout(text_frame)
        text_layout.setContentsMargins(0, 0, 0, 0)
        text_layout.setSpacing(0)
        lbl_name = ClickableLabel(name)
        lbl_name.setFont(kfont(12, True))
        lbl_name.setStyleSheet(f"color:{Palette.text_main};")
        lbl_name.setCursor(Qt.PointingHandCursor)
        lbl_id = ClickableLabel(id_text)
        lbl_id.setFont(kfont(11))
        lbl_id.setStyleSheet(f"color:{Palette.text_sub};")
        lbl_id.setCursor(Qt.PointingHandCursor)
        text_layout.addWidget(lbl_name)
        text_layout.addWidget(lbl_id)
        top_layout.addWidget(text_frame, 1)

        repeat_edit = self._make_repeat_edit()
        top_layout.addWidget(QLabel("반복"))
        top_layout.addWidget(repeat_edit)

        card_layout.addWidget(top_row)

        action_row = QWidget()
        action_layout = QHBoxLayout(action_row)
        action_layout.setContentsMargins(8, 0, 8, 6)
        seg_call = SegmentedButton(seg_call_values, selected_color=Palette.neutral_hover, height=22, font=kfont(10))
        seg_msg = SegmentedButton(seg_msg_values, selected_color=Palette.neutral_hover, height=22, font=kfont(10))
        action_layout.addWidget(seg_call)
        action_layout.addWidget(seg_msg)
        card_layout.addWidget(action_row)
        action_row.setVisible(False)

        def toggle():
            checkbox.setChecked(not checkbox.isChecked())
            checkbox.clicked.emit(checkbox.isChecked())

        lbl_name.clicked.connect(toggle)
        lbl_id.clicked.connect(toggle)
        checkbox.clicked.connect(lambda checked: self._style_checkbox(checkbox))

        parent_layout.insertWidget(parent_layout.count() - 1, card)
        return card, checkbox, seg_call, seg_msg, repeat_edit, action_row

    def _style_checkbox(self, checkbox):
        if checkbox.isChecked():
            checkbox.setIcon(qta.icon("fa5s.check", color="white"))
            checkbox.setIconSize(QSize(11, 11))
            checkbox.setStyleSheet(
                f"QPushButton {{ background-color:{Palette.text_sub}; border:2px solid {Palette.text_sub}; border-radius:3px; }}"
            )
        else:
            checkbox.setIcon(QIcon())
            checkbox.setStyleSheet(
                "QPushButton { background-color:white; border:2px solid #C7C7CC; border-radius:3px; }"
            )

    def refresh_group_list(self):
        if not self.current_uuid:
            return
        path = FileManager.pull_profile_xml(self.current_uuid)
        if not path or not os.path.exists(path):
            return
        groups = FileManager.parse_group_list(path)
        my_info = FileManager.parse_my_info(path)
        self.my_id_label.setText(f"내 정보: {my_info}")

        self.all_cards = []
        self.group_check_vars = {}
        clear_layout(self.group_list_layout, keep=0)
        self.group_list_layout.addStretch(1)

        group_emergency_ok = FileManager.supports_group_emergency_call(self.project_name)
        call_values = ["🔊 PTT", "📹 PTV"]
        if group_emergency_ok:
            call_values += ["🚨 E-PTT", "🚨 E-PTV"]

        def add_section_title(title):
            header = QWidget()
            h_layout = QHBoxLayout(header)
            h_layout.setContentsMargins(4, 10, 4, 4)
            lbl = QLabel(title)
            lbl.setFont(kfont(11, True))
            lbl.setStyleSheet(f"color:{Palette.text_main};")
            h_layout.addWidget(lbl)
            h_layout.addWidget(self._hline(), 1)
            self.group_list_layout.insertWidget(self.group_list_layout.count() - 1, header)

        def add_empty(message):
            lbl = QLabel(message)
            lbl.setFont(kfont(11))
            lbl.setStyleSheet(f"color:{Palette.text_sub}; padding-left:8px;")
            self.group_list_layout.insertWidget(self.group_list_layout.count() - 1, lbl)

        def create_section(title, group_list, empty_message=None):
            if not group_list and not empty_message:
                return
            add_section_title(title)
            if not group_list:
                add_empty(empty_message)
                return

            for g_info in group_list:
                voice, video = g_info.get("voice_codec", ""), g_info.get("video_codec", "")
                codec_str = f"[🎤 {voice} | 🎬 {video}]" if voice or video else ""
                id_text = f"ID: {g_info['id']}"
                if codec_str:
                    id_text += f"  {codec_str}"

                is_emergency = g_info.get("type") == "Emergency"
                if is_emergency:
                    seg_values = ["🚨 Emergency"]
                    if g_info.get("target_type") == "PreArranged Group":
                        seg_values.append("⚠️ Imminent Peril")
                else:
                    seg_values = call_values

                card, checkbox, seg_call, seg_msg, repeat_edit, action_row = self._make_list_card(
                    self.group_list_layout, g_info["name"], id_text, seg_values, ["📄 Text", "🖼️ Photo"]
                )
                checkbox.clicked.connect(self.update_group_visibility)

                check_key = (g_info.get("type"), g_info["id"])
                self.group_check_vars[check_key] = {
                    "id": g_info["id"],
                    "name": g_info["name"],
                    "checkbox": checkbox,
                    "seg_call": seg_call,
                    "seg_msg": seg_msg,
                    "repeat_edit": repeat_edit,
                    "action_row": action_row,
                    "is_emergency": is_emergency,
                }
                self.all_cards.append(self.group_check_vars[check_key])

        def bucket(t):
            return sorted((g for g in groups if g.get("type") == t), key=lambda x: x.get("name", "").lower())

        create_section("🚨 Emergency", bucket("Emergency"), empty_message="설정된 그룹이 없습니다.")
        create_section("📁 ReGroup", bucket("ReGroup"))
        create_section("📁 PreArranged", bucket("PreArranged Group"))
        create_section("💬 Chat", bucket("Chat Group"))
        create_section("💭 Chatting Room", bucket("Chatting"))
        self.update_group_visibility()

    def update_group_visibility(self):
        for data in self.all_cards:
            checkbox = data["checkbox"]
            if not checkbox.isChecked():
                data["seg_call"].set("")
                data["seg_msg"].set("")
                data["action_row"].setVisible(False)
                continue

            data["action_row"].setVisible(True)
            is_call_mode = data.get("is_emergency", False) or self.current_mode == "call"
            data["seg_call"].setVisible(is_call_mode)
            data["seg_msg"].setVisible(not is_call_mode)

    def on_mode_toggle_changed(self, selected_value):
        self.current_mode = "call" if "통화" in selected_value else "msg"
        self.update_group_visibility()
        self.update_user_action_frame()

    def on_main_call_button_clicked(self):
        if not self.current_uuid:
            self.safe_log_insert("⚠️ 단말기가 연결되지 않았습니다!")
            return

        self.stop_event.clear()
        selected_targets = []
        for data in self.group_check_vars.values():
            if not data["checkbox"].isChecked():
                continue
            raw_mode = data["seg_call"].get()
            if not raw_mode:
                self.safe_log_insert(f"⚠️ '{data['name']}' 그룹의 통화 방식이 선택되지 않아 제외됩니다.")
                continue
            clean_mode = raw_mode.split(" ", 1)[-1]
            try:
                repeat_count = int(data["repeat_edit"].text())
            except ValueError:
                repeat_count = 1
            repeat_count = max(1, repeat_count)
            selected_targets.append(
                {"id": data["id"], "name": data["name"], "mode": clean_mode, "repeat": repeat_count}
            )

        if not selected_targets:
            self.safe_log_insert("⚠️ 발신을 진행할 그룹이 없습니다. 체크박스와 통화 방식을 확인해주세요.")
            return

        proj_name = self.project_name
        threading.Thread(
            target=self._process_sequential_calls, args=(proj_name, selected_targets), daemon=True
        ).start()

    def _get_handler(self, proj_name):
        handler_map = {
            "재난망": ("config_handlers.ps_lte_handler", "PsLteHandler"),
            "재난망_LM75": ("config_handlers.ps_lte_lm75_handler", "PsLteLm75Handler"),
            "CTB_POC": ("config_handlers.ctb_poc_handler", "CTB_POCHandler"),
            "450connect": ("config_handlers.connect450_handler", "Connect450Handler"),
        }
        if proj_name not in handler_map:
            return None
        module_name, class_name = handler_map[proj_name]
        module = importlib.import_module(module_name)
        return getattr(module, class_name)()

    def _process_sequential_calls(self, proj_name, selected_targets):
        self.safe_log_insert(f"\n[System] 총 {len(selected_targets)}개 그룹에 순차 발신을 시작합니다...")
        try:
            import uiautomator2 as u2

            d = u2.connect(self.current_uuid)
            handler_instance = self._get_handler(proj_name)
            if handler_instance is None:
                self.safe_log_insert(f"⚠️ '{proj_name}'에 대한 발신 기능이 아직 없습니다.")
                return

            stopped = False
            for idx, target in enumerate(selected_targets, 1):
                if self.stop_event.is_set():
                    stopped = True
                    break
                t_id, t_name, t_mode = target["id"], target["name"], target["mode"]
                t_repeat = target.get("repeat", 1)
                call_target = t_name if proj_name in ("CTB_POC", "450connect") else t_id

                for rep in range(1, t_repeat + 1):
                    if self.stop_event.is_set():
                        stopped = True
                        break
                    self.safe_log_insert(
                        f"\n▶️ [{idx}/{len(selected_targets)}] '{t_name}' ({t_mode}) 발신 진행 중... ({rep}/{t_repeat}회)"
                    )
                    if t_mode in ("Emergency", "Imminent Peril"):
                        if not hasattr(handler_instance, "make_emergency_call"):
                            self.safe_log_insert(f"⚠️ '{proj_name}'에는 비상통화 발신 기능이 아직 없습니다.")
                            continue
                        handler_instance.make_emergency_call(
                            d, imminent=(t_mode == "Imminent Peril"), log_console=self.log_console
                        )
                    else:
                        if not hasattr(handler_instance, "make_call"):
                            self.safe_log_insert(f"⚠️ '{proj_name}'에는 일반 발신 기능이 아직 없습니다.")
                            continue
                        handler_instance.make_call(
                            d, target_info=call_target, call_mode=t_mode, log_console=self.log_console
                        )
                    time.sleep(3)
                if stopped:
                    break

            if stopped:
                self.safe_log_insert("\n⏹ [System] 중지 요청으로 발신을 중단했습니다.")
            else:
                self.safe_log_insert("\n✅ 모든 순차 발신 테스트가 완료되었습니다!")
        except Exception as e:
            self.safe_log_insert(f"❌ 발신 프로세스 중 오류 발생: {e}")

    # ==========================================
    # 👤 User List
    # ==========================================
    def refresh_user_list(self):
        if not self.current_uuid:
            print("❌ 연결된 단말기가 없어 유저 목록을 갱신할 수 없습니다.")
            return

        path = FileManager.pull_profile_xml(self.current_uuid)
        if not path or not os.path.exists(path):
            return

        xml_folder_path = os.path.dirname(os.path.abspath(path))
        users = FileManager.get_all_users_from_xml(xml_folder_path)

        GROUP_CODE_LEN = 3
        my_own_number = FileManager.get_my_own_number(path)
        my_group_code = my_own_number[:GROUP_CODE_LEN] if my_own_number else "006"

        filtered_users = sorted(
            [u for u in users if str(u.get("name", "")).startswith(my_group_code)],
            key=lambda x: str(x.get("name", "")),
        )

        total_found = len(filtered_users)
        MAX_USER_LIST = 30
        if total_found > MAX_USER_LIST:
            filtered_users = filtered_users[:MAX_USER_LIST]
            print(f"👥 [유저 목록 갱신] 총 {total_found}명의 {my_group_code} 유저 중 과부하 방지를 위해 {MAX_USER_LIST}명만 불러옵니다.")
        else:
            print(f"👥 [유저 목록 갱신] 총 {total_found}명의 {my_group_code} 유저를 찾았습니다.")

        clear_layout(self.user_list_layout, keep=0)
        self.user_list_layout.addStretch(1)
        self.user_ui_registry = {}

        for user in filtered_users:
            u_name = user.get("name", "")
            d_name = user.get("display_name", "이름 없음")

            card, checkbox, seg_call, seg_msg, repeat_edit, action_row = self._make_list_card(
                self.user_list_layout,
                d_name,
                f"ID: {u_name}",
                ["🔊 PTT", "📹 PTV", "W/O Floor", "🚨 E-PTT", "🚨 E-PTV"],
                ["📄 Text", "🖼️ Photo", "🎥 Video"],
            )
            checkbox.clicked.connect(self.update_user_action_frame)

            self.user_ui_registry[u_name] = {
                "checkbox": checkbox,
                "seg_call": seg_call,
                "seg_msg": seg_msg,
                "repeat_edit": repeat_edit,
                "action_row": action_row,
            }

    def update_user_action_frame(self):
        for ui_data in self.user_ui_registry.values():
            checkbox = ui_data["checkbox"]
            if not checkbox.isChecked():
                ui_data["seg_call"].set("")
                ui_data["seg_msg"].set("")
                ui_data["action_row"].setVisible(False)
                continue
            ui_data["action_row"].setVisible(True)
            ui_data["seg_call"].setVisible(self.current_mode == "call")
            ui_data["seg_msg"].setVisible(self.current_mode != "call")

    def get_checked_users(self):
        return [u_id for u_id, data in self.user_ui_registry.items() if data["checkbox"].isChecked()]

    def refresh_all_lists(self):
        self.refresh_group_list()
        if self.has_private_call:
            self.refresh_user_list()
        else:
            print("🚫 Private Call 미지원 프로젝트: 유저 데이터 로딩 스킵 (서버 부하 방지)")
            clear_layout(self.user_list_layout, keep=0)
            self.user_list_layout.addStretch(1)
            lbl = QLabel("이 프로젝트는 1:1 통화를 지원하지 않으므로\n유저 목록을 불러오지 않습니다.")
            lbl.setStyleSheet(f"color:{Palette.text_sub};")
            lbl.setFont(kfont(13))
            lbl.setAlignment(Qt.AlignCenter)
            self.user_list_layout.insertWidget(0, lbl)

    def switch_tab(self, tab_name):
        if tab_name == "group":
            self.btn_tab_group.setStyleSheet(btn_css(Palette.neutral_hover, Palette.text_main, Palette.neutral_hover, 17))
            self.btn_tab_user.setStyleSheet(btn_css(Palette.neutral_bg, Palette.text_sub, Palette.neutral_hover, 17))
            self.list_stack.setCurrentIndex(0)
        else:
            self.btn_tab_user.setStyleSheet(btn_css(Palette.neutral_hover, Palette.text_main, Palette.neutral_hover, 17))
            self.btn_tab_group.setStyleSheet(btn_css(Palette.neutral_bg, Palette.text_sub, Palette.neutral_hover, 17))
            self.list_stack.setCurrentIndex(1)

    def send_group_message(self):
        if not self.current_uuid:
            self.safe_log_insert("⚠️ 단말기가 연결되지 않았습니다!")
            return

        self.stop_event.clear()
        selected_groups = []
        for data in self.group_check_vars.values():
            if not data["checkbox"].isChecked():
                continue
            raw_type = data["seg_msg"].get()
            if not raw_type:
                self.safe_log_insert(f"⚠️ '{data['name']}' 그룹의 메시지 방식이 선택되지 않아 제외됩니다.")
                continue
            clean_type = raw_type.split(" ")[-1]
            try:
                repeat_count = int(data["repeat_edit"].text())
            except ValueError:
                repeat_count = 1
            repeat_count = max(1, repeat_count)
            selected_groups.append({"name": data["name"], "msg_type": clean_type, "repeat": repeat_count})

        if not selected_groups:
            self.safe_log_insert("⚠️ 메시지를 보낼 그룹이 없습니다. 체크박스와 메시지 방식을 확인해주세요.")
            return

        proj_name = self.project_name
        threading.Thread(
            target=self._process_sequential_messages, args=(proj_name, selected_groups), daemon=True
        ).start()

    def _process_sequential_messages(self, proj_name, selected_groups):
        self.safe_log_insert(f"\n[System] 총 {len(selected_groups)}개 그룹에 순차 메시지 전송을 시작합니다...")
        try:
            import uiautomator2 as u2

            d = u2.connect(self.current_uuid)
            if proj_name not in ("재난망", "재난망_LM75", "CTB_POC"):
                self.safe_log_insert(f"⚠️ '{proj_name}'에 대한 메시지 전송 기능이 아직 없습니다.")
                return
            handler_instance = self._get_handler(proj_name)

            stopped = False
            for idx, target in enumerate(selected_groups, 1):
                if self.stop_event.is_set():
                    stopped = True
                    break
                t_name, t_type, t_repeat = target["name"], target["msg_type"], target.get("repeat", 1)

                if t_type != "Text":
                    self.safe_log_insert(f"⚠️ '{t_name}' ({t_type}) 메시지 방식은 아직 지원하지 않아 건너뜁니다.")
                    continue

                if proj_name == "CTB_POC":
                    self.safe_log_insert(
                        f"\n▶️ [{idx}/{len(selected_groups)}] '{t_name}' 메시지 전송 진행 중... (총 {t_repeat}회)"
                    )
                    handler_instance.send_message(d, target_info=t_name, repeat=t_repeat, log_console=self.log_console)
                    time.sleep(2)
                else:
                    for rep in range(1, t_repeat + 1):
                        if self.stop_event.is_set():
                            stopped = True
                            break
                        self.safe_log_insert(
                            f"\n▶️ [{idx}/{len(selected_groups)}] '{t_name}' 메시지 전송 진행 중... ({rep}/{t_repeat}회)"
                        )
                        handler_instance.send_message(
                            d, target_info=t_name, seq_no=rep, seq_total=t_repeat, log_console=self.log_console
                        )
                        time.sleep(2)
                    if stopped:
                        break

            if stopped:
                self.safe_log_insert("\n⏹ [System] 중지 요청으로 메시지 전송을 중단했습니다.")
            else:
                self.safe_log_insert("\n✅ 모든 순차 메시지 전송이 완료되었습니다!")
        except Exception as e:
            self.safe_log_insert(f"❌ 메시지 전송 프로세스 중 오류 발생: {e}")

    # ==========================================
    # 📱 미러링
    # ==========================================
    def run_mirror(self):
        self.lbl_placeholder.hide()

        width, height = self.phone_width, self.phone_height
        resolution = adb_logic.get_screen_resolution(self.current_uuid)
        if resolution:
            dev_w, dev_h = resolution
            # 기기 실제 화면 비율에 맞춰 (phone_width, phone_height) 박스 안에 꽉 차도록
            # 축소 비율을 계산합니다. 비율이 정확히 맞아야 스크카피가 여백(레터/필러박스) 없이 채웁니다.
            scale = min(self.phone_width / dev_w, self.phone_height / dev_h)
            width = max(1, round(dev_w * scale))
            height = max(1, round(dev_h * scale))

        self.mirror_container.setFixedSize(width, height)
        parent_hwnd = int(self.mirror_container.winId())
        adb_logic.start_mirroring_embedded(self.current_uuid, parent_hwnd, width, height)

    def record_screen(self):
        self.safe_log_insert("🎥 동영상 촬영 기능 준비 중")

    def capture_screen(self):
        if not self.current_uuid:
            print("⚠️ 연결된 단말기가 없습니다.")
            return
        pictures_dir = os.path.join(os.path.expanduser("~"), "Pictures", "QA_Captures")
        os.makedirs(pictures_dir, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = os.path.join(pictures_dir, f"screenshot_{timestamp}.png")

        print("📸 캡쳐를 진행 중입니다...")
        success = adb_logic.take_screenshot(self.current_uuid, save_path)
        if success:
            print(f"✅ 캡쳐 완료! 파일이 저장되었습니다:\n{save_path}")
            try:
                os.startfile(pictures_dir)
            except Exception:
                pass
        else:
            print("❌ 캡쳐에 실패했습니다.")

    # ==========================================
    # 📦 앱 관리
    # ==========================================
    def run_clear_data(self):
        if not self.current_uuid:
            return
        adb_logic.clear_app_data(self.current_uuid, "com.EveryTalk.Global")
        print("✅ 앱 데이터 초기화 완료")

    def run_install_app(self):
        if not self.current_uuid:
            return
        file_path, _ = QFileDialog.getOpenFileName(self, "설치할 APK 파일을 선택하세요", "", "APK files (*.apk)")
        if not file_path:
            print("사용자가 설치를 취소했습니다.")
            return
        print(f"📂 선택된 파일: {file_path}")
        try:
            print("🚀 설치 진행 중...")
            result = subprocess.run(
                ["adb", "-s", self.current_uuid, "install", "-r", file_path],
                capture_output=True, text=True, check=True,
            )
            if "Success" in result.stdout:
                print("✅ 앱 설치 성공!")
            else:
                print(f"❌ 설치 실패: {result.stderr}")
        except subprocess.CalledProcessError as e:
            print(f"🚨 설치 프로세스 오류: {e.stderr}")
        except Exception as e:
            print(f"🚨 알 수 없는 오류 발생: {e}")

    def run_uninstall_app(self):
        if not self.current_uuid:
            return
        package_name = "com.EveryTalk.Global"
        print(f"🚀 앱 삭제 시도 중: {package_name}")
        try:
            result = subprocess.run(
                ["adb", "-s", self.current_uuid, "uninstall", package_name],
                capture_output=True, text=True,
            )
            if "Success" in result.stdout:
                print("✅ 앱 삭제 성공!")
                self.lbl_version.setText("버전: 삭제됨")
            else:
                print(f"⚠️ 결과: {result.stdout.strip()}")
                if "Failure" in result.stdout:
                    print("❌ 삭제 실패: 앱이 설치되어 있지 않거나 권한이 필요할 수 있습니다.")
        except Exception as e:
            print(f"🚨 삭제 중 에러 발생: {e}")

    # ==========================================
    # ⚙️ 환경 / WiFi 설정 팝업
    # ==========================================
    def _popup_dialog(self, title):
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        dlg.resize(320, 220)
        dlg.setStyleSheet(f"QDialog {{ background-color:{Palette.panel}; }}")
        dlg.setWindowFlags(dlg.windowFlags() | Qt.WindowStaysOnTopHint)
        return dlg

    def open_env_setup(self):
        try:
            with open("env_config.json", "r", encoding="utf-8") as f:
                config_data = json.load(f)
        except FileNotFoundError:
            print("❌ 설정 파일(env_config.json)이 없습니다.")
            return

        project_list = list(config_data.keys())
        dlg = self._popup_dialog("⚙️ 프로젝트 환경 설정")
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(30, 25, 30, 20)

        lbl = QLabel("적용할 프로젝트를 선택하세요")
        lbl.setFont(kfont(14, True))
        lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(lbl)
        layout.addSpacing(15)

        combo = QComboBox()
        combo.addItems(project_list)
        combo.setFixedHeight(36)
        layout.addWidget(combo)
        layout.addSpacing(20)

        btn_apply = self._make_button("✅ 설정 적용", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, height=30)
        btn_apply.clicked.connect(lambda: self.apply_settings(dlg, config_data, combo.currentText()))
        layout.addWidget(btn_apply)

        dlg.show()

    def apply_settings(self, dlg, config_data, proj_name):
        env = config_data[proj_name]
        if self.current_uuid:
            self.safe_log_insert("[System] 자동화 실행 중...")
            try:
                overrides = {
                    "재난망": ("ps_lte", "PsLteHandler"),
                    "재난망_LM75": ("ps_lte_lm75", "PsLteLm75Handler"),
                    "450connect": ("connect450", "Connect450Handler"),
                }
                if proj_name in overrides:
                    safe_proj_name, class_name = overrides[proj_name]
                elif proj_name.lower() == "450connect":
                    safe_proj_name, class_name = "connect450", "Connect450Handler"
                else:
                    safe_proj_name, class_name = proj_name.lower(), f"{proj_name.upper()}Handler"

                module_path = f"config_handlers.{safe_proj_name}_handler"
                module = importlib.import_module(module_path)
                handler = getattr(module, class_name)()

                import uiautomator2 as u2

                d = u2.connect(self.current_uuid)
                handler.run(d, env)

                from common_logger import start_device_logging

                start_device_logging(d, self.log_console)
                self.safe_log_insert("[System] 완료!")
            except ImportError:
                print(f"❌ 모듈 로드 실패: config_handlers.{safe_proj_name}_handler 모듈을 찾을 수 없습니다.")
                self.safe_log_insert("[Error] 설정 실패: 모듈을 찾을 수 없습니다.")
            except Exception as e:
                print(f"❌ 설정 실패: {e}")
                self.safe_log_insert(f"[Error] 설정 실패: {e}")

        dlg.close()

    def open_wifi_setup(self):
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "wifi_config.json")
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                wifi_data = json.load(f)
        except FileNotFoundError:
            print("❌ WiFi 설정 파일(wifi_config.json)이 없습니다.")
            return

        wifi_list = list(wifi_data.keys()) or ["목록 없음"]
        dlg = self._popup_dialog("📶 WiFi 설정")
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(30, 25, 30, 20)

        lbl = QLabel("접속할 WiFi를 선택하세요")
        lbl.setFont(kfont(14, True))
        lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(lbl)
        layout.addSpacing(15)

        combo = QComboBox()
        combo.addItems(wifi_list)
        combo.setFixedHeight(36)
        layout.addWidget(combo)
        layout.addSpacing(20)

        btn_connect_wifi = self._make_button("✅ WiFi 연결", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, height=30)
        btn_connect_wifi.clicked.connect(lambda: self.apply_wifi_settings(dlg, wifi_data, combo.currentText()))
        layout.addWidget(btn_connect_wifi)

        dlg.show()

    def apply_wifi_settings(self, dlg, wifi_data, ssid):
        password = wifi_data.get(ssid)
        if self.current_uuid:
            self.safe_log_insert(f"[System] {ssid} WiFi 연결 시도 중...")
            success = adb_logic.connect_wifi(self.current_uuid, ssid, password)
            if success:
                self.safe_log_insert(f"[System] ✅ {ssid} 연결 성공!")
            else:
                self.safe_log_insert(f"[System] ❌ {ssid} 연결 실패!")
        dlg.close()

    # ==========================================
    # 📝 Logcat / 📡 PCAP
    # ==========================================
    def toggle_log(self):
        if not self.is_log_on:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            log_path = os.path.join(os.getcwd(), "logs", f"log_{self.panel_label}_{timestamp}.txt")
            self.log_proc, self.log_file = adb_logic.start_log_process(self.current_uuid, log_path)
            self.btn_toggle_log.setText("■ LOG OFF")
            self.btn_toggle_log.setStyleSheet(btn_css(Palette.neutral_hover, Palette.text_main, Palette.neutral_hover, 15))
            self.is_log_on = True
        else:
            adb_logic.stop_process(self.log_proc)
            self.log_file.close()
            self.btn_toggle_log.setText("▶ LOG ON")
            self.btn_toggle_log.setStyleSheet(btn_css(Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, 15))
            self.is_log_on = False

    def _set_pcap_ui_state(self, is_on):
        """PCAPdroid 캡처 ON/OFF 상태를 토글 버튼에 반영합니다. 메인 스레드에서만 위젯을
        건드릴 수 있어서, 백그라운드 스레드(자동 셋팅 등)에서는 signals.pcap_state를
        emit해서 이 슬롯을 통해 반영합니다."""
        self.is_pcap_on = is_on
        if is_on:
            self.btn_toggle_pcap.setText("■ PCAPdroid OFF")
            self.btn_toggle_pcap.setStyleSheet(btn_css(Palette.neutral_hover, Palette.text_main, Palette.neutral_hover, 15))
        else:
            self.btn_toggle_pcap.setText("● PCAPdroid ON")
            self.btn_toggle_pcap.setStyleSheet(btn_css(Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, 15))

    def toggle_pcap(self):
        if not self.current_uuid:
            self.safe_log_insert("[System] ❌ 먼저 단말기를 연결해 주세요.")
            return

        if not self.is_pcap_on:
            self.safe_log_insert("[System] PCAPdroid 상태 점검 및 실행 중...")
            self.stop_realtime_sip_stream()

            def log(msg):
                self.safe_log_insert(f"[PCAPdroid] {msg}", is_error=("❌" in msg))

            adb_logic.switch_pcapdroid_to_pcap_file_mode(self.current_uuid, log=log)
            success = adb_logic.start_pcapdroid(self.current_uuid, log=log)
            if success:
                self._set_pcap_ui_state(True)
                self.safe_log_insert("[System] 📡 PCAPdroid 캡처 활성화 완료!")
            else:
                self.safe_log_insert("[System] ❌ 캡처를 시작하지 못했습니다. 로그를 확인하세요.")
        else:
            self.safe_log_insert("[System] 캡처 종료 중...")

            def log(msg):
                self.safe_log_insert(f"[PCAPdroid] {msg}", is_error=("❌" in msg))

            adb_logic.stop_pcapdroid(self.current_uuid, log=log)
            self._set_pcap_ui_state(False)
            self.safe_log_insert("[System] 🛑 캡처가 중지되었습니다.")

            threading.Thread(target=self._analyze_pcap_sip_flow, args=(self.current_uuid,), daemon=True).start()
            self.start_realtime_sip_stream()

    def _analyze_pcap_sip_flow(self, uuid):
        local_dir = os.path.join(os.getcwd(), "pcap_captures")
        self.safe_log_insert("\n[System] 📥 pcap 파일을 폰에서 가져오는 중...")

        pcap_path = adb_logic.pull_latest_pcapdroid_file(uuid, local_dir)
        if not pcap_path:
            self.safe_log_insert("[System] ❌ pcap 파일을 가져오지 못해 SIP Flow 분석을 건너뜁니다.")
            return

        self.safe_log_insert(f"[System] 🔍 tshark로 SIP 메시지를 분석 중입니다: {os.path.basename(pcap_path)}")
        events = adb_logic.parse_sip_flow_from_pcap(pcap_path)
        if not events:
            self.safe_log_insert("[System] ⚠️ pcap에서 SIP 메시지를 찾지 못했습니다 (tshark 설치 여부를 확인해주세요).")
            return

        self.add_flow_card("PROC", "PCAP 분석 완료", f"'{os.path.basename(pcap_path)}'에서 SIP 메시지 {len(events)}건을 찾았습니다.")
        for ev in events:
            event_type = "RX" if ev["is_response"] else "PROC"
            self.add_flow_card(event_type, ev["title"], ev["detail"])

        self.safe_log_insert("[System] ✅ SIP Flow(패킷 기반) 분석 완료!")

    # ==========================================
    # 📡 실시간 SIP Flow (logcat 기반 - 네트워크 캡처 없음)
    # ==========================================
    def start_realtime_sip_stream(self):
        if not self.current_uuid or self.is_pcap_on:
            return
        if self._realtime_sip_thread and self._realtime_sip_thread.is_alive():
            return

        uuid = self.current_uuid
        stop_event = threading.Event()
        state = {}
        self._realtime_sip_stop_event = stop_event
        self._realtime_sip_state = state

        def on_event(ev):
            event_type = "RX" if ev["is_response"] else "PROC"
            self.add_flow_card(event_type, ev["title"], ev["detail"])

        def log(msg):
            self.safe_log_insert(f"[SIP Flow] {msg}", is_error=("❌" in msg))

        def worker():
            adb_logic.run_logcat_sip_stream(uuid, on_event, stop_event, state, log=log)

        self._realtime_sip_thread = threading.Thread(target=worker, daemon=True)
        self._realtime_sip_thread.start()

    def stop_realtime_sip_stream(self):
        stop_event = self._realtime_sip_stop_event
        state = self._realtime_sip_state
        if stop_event:
            stop_event.set()
        if state:
            for key in ("proc", "conn", "server"):
                obj = state.get(key)
                if not obj:
                    continue
                try:
                    obj.terminate() if key == "proc" else obj.close()
                except Exception:
                    pass
        self._realtime_sip_stop_event = None
        self._realtime_sip_state = None

    # ==========================================
    # 📡 실시간 SIP/Call Flow 로그 분석
    # ==========================================
    def _emit_flow(self, key, event_type, title, detail, is_error=False, window=1.5):
        """같은 종류(key)의 이벤트가 짧은 시간 안에 같은 내용으로 반복되면 무시합니다.
        EveryTalk 앱은 같은 이벤트를 여러 UI 컴포넌트에서 각각 로깅해서 실제로
        한 번의 통화에 동일 라인이 6~8번씩 찍히는 걸 실측으로 확인했습니다."""
        now = time.time()
        last = self._flow_dedupe.get(key)
        if last and last[0] == detail and (now - last[1]) < window:
            self._flow_dedupe[key] = (detail, now)
            return False
        self._flow_dedupe[key] = (detail, now)
        self.add_flow_card(event_type, title, detail, is_error=is_error)
        return True

    def start_realtime_log_analyzer(self):
        if not self.current_uuid:
            return

        if self._sip_log_process:
            try:
                self._sip_log_process.terminate()
            except Exception:
                pass

        self._sip_analyzer_gen += 1
        gen = self._sip_analyzer_gen
        uuid = self.current_uuid
        self._flow_dedupe = {}
        self._pending_dnd_reason = False
        self._show_sip_placeholder()

        def logcat_reader():
            subprocess.run(["adb", "-s", uuid, "logcat", "-c"])
            process = subprocess.Popen(
                ["adb", "-s", uuid, "logcat", "-v", "time"],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
            self._sip_log_process = process
            self.add_flow_card("PROC", "Analyzer Started", "실시간 SIP/Call Flow 감지를 시작합니다.")

            while self._sip_analyzer_gen == gen:
                line = process.stdout.readline()
                if not line:
                    break

                is_relevant_error = ("Exception" in line or " E " in line) and (
                    "MCPTT" in line or "EveryTalk" in line
                )
                if "Exception" in line:
                    self.safe_log_insert(line, is_error=is_relevant_error)

                if "getCalleeUri: calleeUri =" in line:
                    target = line.split("calleeUri =")[1].strip()
                    self._emit_flow("callee", "PROC", "Call Target", f"대상: {target}")
                    self.safe_log_insert(line)
                elif "uiEventType = TYPE_REQUEST_CALL" in line:
                    self._emit_flow("call_state", "PROC", "Call Requested", "발신 요청")
                    self.safe_log_insert(line)
                elif "uiEventType = TYPE_MEDIA_PREPARE_COMPLETE" in line:
                    self._emit_flow("call_state", "PROC", "Media Ready", "미디어 준비 완료")
                    self.safe_log_insert(line)
                elif "uiEventType = TYPE_CALL_CONNECT_OK" in line:
                    self._pending_dnd_reason = False
                    self._emit_flow("call_state", "RX", "Call Connected", "통화 연결 성공")
                    if self.results_panel:
                        self.results_panel.auto_grade_result("PASS", f"[{self.panel_label}] 통화 연결 성공 (자동 감지)")
                    self.safe_log_insert(line)
                elif "uiEventType = TYPE_INFO_CALL_DND" in line:
                    self._pending_dnd_reason = True
                    self.safe_log_insert(line)
                elif self._pending_dnd_reason and "extra = " in line:
                    self._pending_dnd_reason = False
                    reason = line.split("extra = ", 1)[1].strip()
                    if reason and reason != "null":
                        self._emit_flow("call_fail", "ERR", "Call Failed", reason, is_error=True)
                        if self.results_panel:
                            self.results_panel.auto_grade_result("FAIL", f"[{self.panel_label}] 통화 실패 자동 감지: {reason}")
                    self.safe_log_insert(line)
                elif "uiEventType = TYPE_DELETE_SESSION" in line or "uiEventType = TYPE_DELETED_SESSION" in line:
                    self._emit_flow("session_end", "PROC", "Session End", "세션 종료")
                    self.safe_log_insert(line)
                elif "mMBCPKeyEvent =" in line:
                    state = line.split("mMBCPKeyEvent =")[1].strip()
                    changed = self._emit_flow("mbcp", "RX", "Floor State", f"상태: {state}")
                    if changed:
                        self.signals.floor_state.emit(state)
                    self.safe_log_insert(line)
                elif is_relevant_error and "Exception" not in line:
                    self.safe_log_insert(line, is_error=True)

            if self.current_uuid == uuid:
                process.terminate()

        threading.Thread(target=logcat_reader, daemon=True).start()

    # ==========================================
    # 🧪 단위 테스트 팝업 / 자동화
    # ==========================================
    def open_unit_test_popup(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("단위 테스트 시나리오")
        dlg.resize(300, 450)
        dlg.setWindowFlags(dlg.windowFlags() | Qt.WindowStaysOnTopHint)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(15, 15, 15, 15)

        scroll, container, content_layout = self._make_scroll_list()
        layout.addWidget(scroll)

        menu_data = {
            "📞 Group Call": ["ReGroup", "PreArranged", "Chat Group"],
            "👤 Private Call": ["Private PTT", "Private PTV", "MCVideo Push"],
            "💬 IM Message": ["일반 메시지", "사진 첨부", "동영상", "기타문서"],
        }
        for category, items in menu_data.items():
            lbl = QLabel(category)
            lbl.setFont(kfont(13, True))
            lbl.setStyleSheet(f"color:{Palette.blue};")
            content_layout.insertWidget(content_layout.count() - 1, lbl)
            for item in items:
                btn = self._make_button(f"  {item}", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover, height=28, radius=4)
                btn.setStyleSheet(btn.styleSheet() + "QPushButton { text-align:left; padding-left:10px; }")
                btn.clicked.connect(lambda checked=False, c=category, i=item: self.execute_action(c, i))
                content_layout.insertWidget(content_layout.count() - 1, btn)

        dlg.show()

    def execute_action(self, category, item):
        print(f"▶ 개별 단위 테스트 실행: [{category}] -> {item}")

    def run_automation(self):
        if not self.current_uuid:
            return
        print("🚀 [자동화 시작] 시나리오를 연속 실행합니다...")

    def stop_automation(self):
        print("⏹ [자동화 중지] 시나리오를 강제 중지합니다.")
        self.stop_event.set()
        self.safe_log_insert("\n⏹ [System] 중지 요청됨. 진행 중인 항목을 마치는 대로 멈춥니다.")


class ResultsPanel(QWidget):
    """Test Results 표(엑셀 시트 형식). 결국 하나의 TC 목록을 두 단말이 같이 채워나가는
    것이므로, 패널마다 따로 두지 않고 App이 이 위젯을 하나만 만들어서 두 DevicePanel이
    공유합니다 (양쪽 DevicePanel 모두 이 인스턴스를 self.results_panel로 들고 있습니다).
    화면에는 App(ui_logic.py)이 두 패널의 로그카드(Network & System Logs) 아래에
    두 로그카드 폭만큼만 걸치는 하나의 병합된 띠로 배치합니다(미러링 화면 아래까지는
    내려오지 않습니다)."""

    # 통화 연결/거부 이벤트를 감지했을 때, 이미 가져와둔 TC 목록의 '요약' 칸에서
    # 이 단어들을 찾아 자동으로 PASS/FAIL을 채웁니다. 실제 테스트케이스 문구에 맞게
    # 이 목록만 수정하면 됩니다.
    AUTO_PASS_KEYWORDS = ("수신", "착신", "연결")
    AUTO_FAIL_KEYWORDS = ("실패", "거부", "안됨", "불가")

    def __init__(self, parent=None):
        super().__init__(parent)
        styled(self, card_css())
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(6)

        header = QHBoxLayout()
        lbl_title = QLabel("Test Results")
        lbl_title.setFont(kfont(13, True))
        lbl_title.setStyleSheet(f"color:{Palette.text_main};")
        header.addWidget(lbl_title)
        header.addStretch(1)
        self.btn_import_results = make_button(
            "가져오기", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover,
            height=24, radius=5, icon_name="fa5s.file-import",
        )
        self.btn_import_results.setFixedWidth(90)
        self.btn_import_results.clicked.connect(self.import_results_from_file)
        header.addWidget(self.btn_import_results)
        self.btn_download_results = make_button(
            "다운로드", Palette.neutral_bg, Palette.text_main, Palette.neutral_hover,
            height=24, radius=5, icon_name="fa5s.download",
        )
        self.btn_download_results.setFixedWidth(100)
        self.btn_download_results.clicked.connect(self.export_results_to_excel)
        header.addWidget(self.btn_download_results)
        layout.addLayout(header)

        self.table_results = QTableWidget(0, 3)
        self.table_results.setHorizontalHeaderLabels(["TC-ID", "요약", "결과"])
        self.table_results.verticalHeader().setVisible(False)
        self.table_results.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_results.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_results.setAlternatingRowColors(False)
        self.table_results.setShowGrid(True)
        excel_grid_color = "#D4D4D4"
        # 엑셀 기본 행 높이 15.75pt -> px (96dpi 기준: pt * 96/72)
        excel_row_height = round(15.75 * 96 / 72)
        header_view = self.table_results.horizontalHeader()
        header_view.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header_view.setSectionResizeMode(1, QHeaderView.Stretch)
        header_view.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        v_header = self.table_results.verticalHeader()
        v_header.setDefaultSectionSize(excel_row_height)
        v_header.setMinimumSectionSize(excel_row_height)
        v_header.setSectionResizeMode(QHeaderView.Fixed)
        self.table_results.setStyleSheet(
            "QTableWidget {"
            f"  background-color:white; border:1px solid {excel_grid_color}; border-radius:0px;"
            f"  gridline-color:{excel_grid_color}; color:{Palette.text_main};"
            "}"
            "QTableWidget::item {"
            "  padding:2px 8px; background-color:white;"
            f"  border-right:1px solid {excel_grid_color}; border-bottom:1px solid {excel_grid_color};"
            "}"
            f"QTableWidget::item:selected {{ background-color:{Palette.tint_blue_bg}; color:{Palette.text_main}; }}"
            "QHeaderView::section {"
            f"  background-color:{Palette.neutral_bg}; color:{Palette.text_main}; font-weight:600;"
            "  border:none;"
            "  padding:2px 8px;"
            "}"
        )
        layout.addWidget(self.table_results, 1)

        # 데이터가 없어도 엑셀 시트처럼 기본 격자선이 보이도록 빈 행을 미리 채워둡니다.
        self._add_blank_result_rows(20)

    def _add_blank_result_rows(self, count):
        start = self.table_results.rowCount()
        self.table_results.setRowCount(start + count)
        for row in range(start, start + count):
            for col in range(3):
                self.table_results.setItem(row, col, QTableWidgetItem(""))

    def _next_blank_result_row(self):
        for row in range(self.table_results.rowCount()):
            item = self.table_results.item(row, 0)
            if item is None or not item.text():
                return row
        return None

    def _make_result_item(self, result):
        item = QTableWidgetItem(str(result))
        if "실패" in str(result) or "FAIL" in str(result).upper():
            item.setForeground(QColor(Palette.danger))
        elif "성공" in str(result) or "PASS" in str(result).upper():
            item.setForeground(QColor(Palette.blue))
        return item

    def add_result_row(self, tc_id, summary, result):
        """TC-ID / 요약 / 결과 한 줄을 결과 테이블에 추가합니다."""
        row = self._next_blank_result_row()
        if row is None:
            row = self.table_results.rowCount()
            self._add_blank_result_rows(1)
        self.table_results.setItem(row, 0, QTableWidgetItem(str(tc_id)))
        self.table_results.setItem(row, 1, QTableWidgetItem(str(summary)))
        self.table_results.setItem(row, 2, self._make_result_item(result))

    def auto_grade_result(self, verdict, detail=""):
        """단말A/단말B 어느 쪽 로그에서든 통화 연결/거부 이벤트를 감지하면 호출됩니다.
        이미 가져와둔(또는 직접 입력한) TC 목록 중 '요약'에 관련 키워드가 있고 '결과'가
        아직 비어있는 첫 번째 행을 찾아 PASS/FAIL을 채우고, 해당하는 TC가 없으면
        (또는 있던 매칭 행들이 이미 다 채점된 상태면) 새 행을 추가합니다."""
        keywords = self.AUTO_PASS_KEYWORDS if verdict == "PASS" else self.AUTO_FAIL_KEYWORDS

        target_row = None
        for row in range(self.table_results.rowCount()):
            summary_item = self.table_results.item(row, 1)
            if summary_item is None or not summary_item.text().strip():
                continue
            result_item = self.table_results.item(row, 2)
            if result_item is not None and result_item.text().strip():
                continue
            if any(kw in summary_item.text() for kw in keywords):
                target_row = row
                break

        if target_row is not None:
            self.table_results.setItem(target_row, 2, self._make_result_item(verdict))
            tc_item = self.table_results.item(target_row, 0)
            tc_id = tc_item.text() if tc_item else ""
            print(f"[Results] ✅ 자동 채점: '{tc_id}' 결과에 {verdict} 반영 (감지: {detail or verdict})")
        else:
            timestamp = datetime.datetime.now().strftime("%H%M%S")
            auto_tc_id = f"AUTO-{timestamp}"
            summary = detail or ("통화 연결 자동 감지" if verdict == "PASS" else "통화 실패 자동 감지")
            self.add_result_row(auto_tc_id, summary, verdict)
            print(f"[Results] ✅ 자동 채점: 일치하는 TC가 없어 새 행 추가 ('{auto_tc_id}') → {verdict}")

    def export_results_to_excel(self):
        """결과 테이블을 엑셀(csv) 파일로 내보냅니다."""
        rows = []
        for row in range(self.table_results.rowCount()):
            tc_item = self.table_results.item(row, 0)
            if tc_item is None or not tc_item.text():
                continue
            rows.append([
                self.table_results.item(row, col).text() if self.table_results.item(row, col) else ""
                for col in range(3)
            ])

        if not rows:
            QMessageBox.information(self, "내보내기", "내보낼 테스트 결과가 없습니다.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "테스트 결과 저장", "test_results.csv", "CSV Files (*.csv)"
        )
        if not path:
            return

        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["TC-ID", "요약", "결과"])
            writer.writerows(rows)
        print(f"[Results] 테스트 결과를 저장했습니다: {path}")

    def import_results_from_file(self):
        """엑셀(.xlsx) 또는 CSV로 정리해둔 테스트케이스 목록(TC-ID, 요약)을 읽어와
        결과표를 미리 채워둡니다. '결과' 칸은 비워두므로, 테스트를 진행하면서 그 칸만
        채우면 됩니다. 기존에 입력해둔 내용은 덮어씌워지니 주의가 필요합니다."""
        path, _ = QFileDialog.getOpenFileName(
            self, "테스트케이스 목록 가져오기", "", "Excel/CSV Files (*.xlsx *.csv);;Excel (*.xlsx);;CSV (*.csv)"
        )
        if not path:
            return

        try:
            rows = self._read_case_rows(path)
        except Exception as e:
            print(f"[Results] 결과 가져오기 실패: {e!r}")
            QMessageBox.warning(self, "가져오기 실패", f"파일을 읽는 중 오류가 발생했습니다.\n\n{e}")
            return

        if not rows:
            msg = "파일에서 테스트케이스를 찾지 못했습니다.\n1열=TC-ID, 2열=요약 형식인지 확인해주세요."
            QMessageBox.warning(self, "가져오기 실패", msg)
            return

        self.table_results.setRowCount(0)
        for tc_id, summary in rows:
            row = self.table_results.rowCount()
            self.table_results.insertRow(row)
            self.table_results.setItem(row, 0, QTableWidgetItem(tc_id))
            self.table_results.setItem(row, 1, QTableWidgetItem(summary))
            self.table_results.setItem(row, 2, QTableWidgetItem(""))
        self._add_blank_result_rows(10)
        print(f"[Results] ✅ 테스트케이스 {len(rows)}건을 가져왔습니다: {os.path.basename(path)}")
        QMessageBox.information(self, "가져오기 완료", f"테스트케이스 {len(rows)}건을 가져왔습니다.")

    def _read_case_rows(self, path):
        """CSV 또는 엑셀(.xlsx) 파일에서 (TC-ID, 요약) 튜플 목록을 읽어옵니다.
        첫 줄이 'TC-ID'/'요약' 같은 헤더로 보이면 건너뜁니다."""
        ext = os.path.splitext(path)[1].lower()
        raw_rows = []
        if ext == ".csv":
            import csv
            with open(path, "r", newline="", encoding="utf-8-sig") as f:
                for r in csv.reader(f):
                    raw_rows.append(r)
        elif ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                raw_rows.append(["" if v is None else str(v) for v in r])
        else:
            raise ValueError(f"지원하지 않는 파일 형식입니다: {ext} (.xlsx 또는 .csv만 가능합니다. 오래된 .xls는 .xlsx로 다시 저장해주세요.)")

        rows = []
        for r in raw_rows:
            if not r or not str(r[0]).strip():
                continue
            tc_id = str(r[0]).strip()
            summary = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
            if self._looks_like_header_row(tc_id, summary):
                continue
            rows.append((tc_id, summary))
        return rows

    def _looks_like_header_row(self, tc_id, summary):
        """"TC-ID"/"TC ID"/"tc_id"처럼 공백·하이픈·대소문자가 달라도 헤더 행으로
        인식하도록, 문자만 남기고 대문자로 정규화해서 비교합니다."""
        norm_id = re.sub(r"[^0-9A-Za-z가-힣]", "", tc_id).upper()
        norm_summary = re.sub(r"[^0-9A-Za-z가-힣]", "", summary).upper()
        return norm_id in ("TCID", "ID", "NO", "번호") or norm_summary in ("요약", "SUMMARY", "설명", "DESCRIPTION")
